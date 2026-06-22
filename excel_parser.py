import io
from datetime import date, datetime, timezone, timedelta
import openpyxl

def _hoje_brt():
    return datetime.now(timezone(timedelta(hours=-3))).date()

def _parse_dt(val):
    if not val:
        return '', 0
    hoje = _hoje_brt()
    # openpyxl pode retornar datetime ou date diretamente
    if isinstance(val, datetime):
        d = val.date()
        return d.strftime('%d/%m/%Y'), (hoje - d).days
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y'), (hoje - val).days
    s = str(val).strip()
    if not s or s == 'None':
        return '', 0
    s_date = s[:10]
    # tenta vários formatos incluindo traço (DD-MM-YYYY) e barra (DD/MM/YYYY)
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%m-%d-%Y'):
        try:
            d = datetime.strptime(s_date, fmt).date()
            return d.strftime('%d/%m/%Y'), (hoje - d).days
        except ValueError:
            continue
    return s_date, 0

def _bool_val(val):
    if val is None:
        return False
    return str(val).strip().lower() in ('sim', 'true', '1', 'yes', 's')

def parse_excel(content_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []

    hdr = {}
    for i, v in enumerate(rows[0]):
        if v is not None:
            hdr[str(v).strip()] = i

    def get(row, col, fb=''):
        idx = hdr.get(col)
        if idx is None or idx >= len(row):
            return fb
        v = row[idx]
        return str(v).strip() if v is not None else fb

    def get_raw(row, col):
        idx = hdr.get(col)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    STATUS_OPEN   = {'Novo', 'Em andamento', 'Aguardando'}
    STATUS_CLOSED = {'Resolvido', 'Fechado', 'Cancelado'}
    hoje = _hoje_brt()

    main_tks = []
    baixados = []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        if not code.isdigit():
            continue

        status = get(row, 'status')

        data_str, dias      = _parse_dt(get_raw(row, 'datahora_abertura'))
        raw_mov             = get_raw(row, 'datahora_ultima_interacao')
        _,        dias_mov  = _parse_dt(raw_mov)
        # se não tem data de movimentação, considera que nunca foi tocado (usa dias em aberto)
        if not raw_mov or str(raw_mov).strip() in ('', 'None'):
            dias_mov = max(dias, 0)
        resolucao, _        = _parse_dt(get_raw(row, 'datahora_encerramento'))

        t = {
            'code':         code,
            'produto':      get(row, 'produtoId'),
            'status':       status,
            'tipo':         get(row, 'tipoId'),
            'prioridade':   '',
            'empresa':      get(row, 'empresaId') or 'SEM EMPRESA',
            'grupo':        get(row, 'grupo'),
            'atrib':        get(row, 'atribuido') or '— Sem Responsável —',
            'assunto':      get(row, 'assunto'),
            'data':         data_str,
            'dias':         max(dias, 0),
            'dias_sem_mov': max(dias_mov, 0),
            'sla_venceu':   _bool_val(get(row, 'sla_info.venceu_resolucao')),
            'abandonado':   _bool_val(get(row, 'abandonado')),
            'resolucao':    resolucao,
            'ordem':        get(row, 'ordem-logussul'),
        }

        if status in STATUS_OPEN:
            main_tks.append(t)
        elif status in STATUS_CLOSED and resolucao:
            dt_enc = get_raw(row, 'datahora_encerramento')
            _, days_ago = _parse_dt(dt_enc)
            if days_ago == 0:
                baixados.append(t)

    return main_tks, baixados


def parse_backlog_excel(conteudo):
    """Lê o arquivo BACKLOG CUSTOMIZAÇÕES e retorna lista no formato do BACKLOG."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active
    backlog = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        num    = row[0]
        ticket = row[2]
        cliente= row[3]
        desc   = row[4]
        est_v  = row[5]
        status = row[6]
        # pula linhas sem ticket numérico
        if not ticket or not str(ticket).replace('.0','').strip().isdigit():
            continue
        num = int(num) if num else len(backlog) + 1
        ticket_str = str(int(ticket))
        est = ''
        if isinstance(est_v, datetime):
            est = est_v.strftime('%d/%m')
        elif est_v:
            est = str(est_v).strip()
        backlog.append({
            'num':    num,
            'ticket': ticket_str,
            'cliente': str(cliente).strip() if cliente else '',
            'desc':   str(desc).strip()    if desc    else '',
            'est':    est,
            'status': str(status).strip()  if status  else 'Pendente',
        })
    return backlog
